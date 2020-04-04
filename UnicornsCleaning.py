import openpyxl as xl

path = "Unicorns_Dataset.xlsx"

wb = xl.load_workbook(path)
ws = wb["Investors"]
ws2 = wb.create_sheet("Cleaned", index=4)


def create_list(ws, column):
    new_list = []
    for i in range(1, ws.max_row + 1):
        company = ws.cell(row=i, column=column).value
        new_list.append(company)
    return new_list


def check_company(list1, list2):
    for company in list2:
        if company in list1:
            continue
        else:
            list1.append(company)
    return list1


def add_to_worksheet(list_name, ws, row, column):
    for company in list_name:
        ws.cell(row=row, column=column).value = company
        row += 1


col1 = create_list(ws, 1)
col2 = create_list(ws, 2)
col3 = create_list(ws, 3)
col4 = create_list(ws, 4)

check_company(col1, col2)
check_company(col1, col3)
check_company(col1, col4)

add_to_worksheet(col1, ws2, 1, 1)
wb.save(path)
