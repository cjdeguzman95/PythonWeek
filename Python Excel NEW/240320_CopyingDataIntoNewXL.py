import xlsxwriter as xw
import openpyxl as xl

# create new destination excel file
wb2 = xw.Workbook("Data 9 Copy.xlsx")
ws2 = wb2.add_worksheet("Trainees")
wb2.close()

# open original file
path = "Data 9.xlsx"
wb1 = xl.load_workbook(path)
ws1 = wb1.worksheets[0]

# open destination file
path2 = "Data 9 Copy.xlsx"
wb2 = xl.load_workbook(path2)
ws2 = wb2.worksheets[0]

# copying values and pasting into new file
for row in range(1, ws1.max_row + 1):
    for column in range(1, ws1.max_column + 1):
        cell_value = ws1.cell(row=row, column=column)
        ws2.cell(row=row, column=column).value = cell_value.value


wb2.save(path2)
