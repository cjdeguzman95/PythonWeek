# change the title and a cell value in an existing excel file
import openpyxl as xl

path = "Data 9.xlsx"

workbook = xl.load_workbook(path)
worksheet = workbook["Trainees"]
worksheet = workbook.active

# this changes the title of the worksheet from Trainees to Students
worksheet.title = "Students"

# this adds trainee id into Column A (previously missing from the worksheet)
row = 3

for i in range(3, worksheet.max_row + 1):
    worksheet.cell(row=row, column=1).value = i
    row += 1

# change the name 'Gigi' to 'Wan Chi' = B5
changed_cell = worksheet.cell(row=5, column=2).value = "Wan Chi"

# displays the values in the worksheet on the terminal
for row in range(1, worksheet.max_row + 1):
    for column in range(1, worksheet.max_column + 1):
        print(worksheet.cell(row, column).value, end=" ")
    print()

print("No. of rows: {}".format(worksheet.max_row))
print("No. of columns: {}".format(worksheet.max_column))
workbook.save(path)
